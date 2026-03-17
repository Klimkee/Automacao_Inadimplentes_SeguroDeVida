from __future__ import annotations
import openpyxl 
import argparse
import os
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FINAL_COLUMNS = [
    "Nome segurado",
    "Data contato",
    "Número apólice",
    "Account ID",
    "Seguradora",
    "Dias atraso",
    "Data vencimento",
    "Valor parcela",
    "Total inadimplente",
    "Periodicidade",
    "Método de pagamento",
    "Proprietário da Oportunidade",
    "E-mail do Proprietário da Oportunidade",
    "Business Channel",
]

BASE_COLUMNS = [
    "Nome segurado",
    "Data contato",
    "Número apólice",
    "Seguradora",
    "Dias atraso",
    "Data vencimento",
    "Valor parcela",
    "Total inadimplente",
]

ENRICH_COLUMNS = [
    "Account ID",
    "Periodicidade",
    "Método de pagamento",
    "Proprietário da Oportunidade",
    "E-mail do Proprietário da Oportunidade",
    "Business Channel",
]

FLOW_ORDER = [
    "azos",
    "mag",
    "omint",
    "icatu",
    "metlife",
    "prudential",
]

INSURER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "azos": ("azos",),
    "mag": ("mag",),
    "omint": ("omint",),
    "icatu": ("icatu",),
    "metlife": ("metlife", "met life"),
    "prudential": ("prudential", "pru"),
}

BASE_SF_FILE_NAME = "BASE_SF_INADIMPLENCIA.xlsx"
BASE_SF_SHEET_NAME = "Consulta1"


@dataclass(frozen=True)
class LayoutConfig:
    seguradora: str
    keep_letters: List[str]
    source_to_target: Dict[str, str]


@dataclass
class LookupLoadResult:
    df: pd.DataFrame
    warnings: List[str]


CONFIGS: Dict[str, LayoutConfig] = {
    "azos": LayoutConfig(
        seguradora="Azos",
        keep_letters=["A", "I", "L", "Q"],
        source_to_target={
            "A": "Nome segurado",
            "I": "Número apólice",
            "L": "Valor parcela",
            "Q": "Data vencimento",
        },
    ),
    "mag": LayoutConfig(
        seguradora="MAG",
        keep_letters=["D", "I", "W", "AB"],
        source_to_target={
            "D": "Nome segurado",
            "I": "Número apólice",
            "W": "Data vencimento",
            "AB": "Valor parcela",
        },
    ),
    "omint": LayoutConfig(
        seguradora="Omint",
        keep_letters=["B", "F", "I", "K"],
        source_to_target={
            "B": "Nome segurado",
            "F": "Número apólice",
            "I": "Data vencimento",
            "K": "Valor parcela",
        },
    ),
    "metlife": LayoutConfig(
        seguradora="MetLife",
        keep_letters=["A", "C", "L", "M"],
        source_to_target={
            "A": "Número apólice",
            "C": "Nome segurado",
            "L": "Valor parcela",
            "M": "Data vencimento",
        },
    ),
    "icatu": LayoutConfig(
        seguradora="Icatu",
        keep_letters=[],
        source_to_target={},
    ),
    "prudential": LayoutConfig(
        seguradora="Prudential",
        keep_letters=["E", "G", "Q", "S"],
        source_to_target={
            "E": "Número apólice",
            "G": "Nome segurado",
            "Q": "Data vencimento",
            "S": "Valor parcela",
        },
    ),
}


def excel_letter_to_index(letter: str) -> int:
    result = 0
    for ch in letter.upper().strip():
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Letra de coluna inválida: {letter}")
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def normalize_policy_number(value: object) -> object:
    if value is None or pd.isna(value):
        return pd.NA

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if pd.isna(value):
            return pd.NA
        return f"{value:.0f}"

    text = str(value).strip()
    if not text:
        return pd.NA

    scientific = text.replace(",", ".")
    if "e" in scientific.lower():
        try:
            return f"{float(scientific):.0f}"
        except ValueError:
            pass

    digits = re.sub(r"\D", "", text)
    return digits if digits else pd.NA


def metlife_prefix_policy(value: object) -> object:
    normalized = normalize_policy_number(value)
    if pd.isna(normalized):
        return pd.NA

    s = re.sub(r"\D", "", str(normalized))
    if not s:
        return pd.NA

    if len(s) == 5:
        return "9100" + s
    if len(s) == 6:
        return "910" + s

    return s


def normalize_brl_number(value: object) -> object:
    if value is None or pd.isna(value):
        return pd.NA

    if isinstance(value, (int, float, Decimal)):
        return float(value)

    text = str(value).strip()
    if not text:
        return pd.NA

    text = text.replace("R$", "").replace(" ", "")
    text = re.sub(r"[^0-9,.\-]", "", text)

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    

    try:
        return float(Decimal(text))
    except (InvalidOperation, ValueError):
        return pd.NA


def distinct_non_empty_values(series: pd.Series) -> List[str]:
    values: List[str] = []
    seen = set()

    for value in series:
        if value is None or pd.isna(value):
            continue

        text = str(value).strip()
        if not text or text in seen:
            continue

        seen.add(text)
        values.append(text)

    return values


def is_layout_configured(config: LayoutConfig) -> bool:
    return bool(config.keep_letters and config.source_to_target)


HEADER_FILL = PatternFill("solid", fgColor="184E57")
ALT_FILL = PatternFill("solid", fgColor="F6F1E7")
CARD_FILL = PatternFill("solid", fgColor="0F2D33")
ACCENT_FILL = PatternFill("solid", fgColor="C48A2A")
SOFT_FILL = PatternFill("solid", fgColor="E7F1EF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="184E57", bold=True, size=18)
SUBTITLE_FONT = Font(color="4B5F63", italic=True)
CARD_TITLE_FONT = Font(color="FFFFFF", bold=True, size=10)
CARD_VALUE_FONT = Font(color="FFFFFF", bold=True, size=16)
SECTION_FONT = Font(color="184E57", bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style="thin", color="D8E0DF"),
    right=Side(style="thin", color="D8E0DF"),
    top=Side(style="thin", color="D8E0DF"),
    bottom=Side(style="thin", color="D8E0DF"),
)


def format_brl(value: float) -> str:
    formatted = f"{value:,.2f}"
    return "R$ " + formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def auto_fit_columns(ws, df: pd.DataFrame) -> None:
    if df.empty:
        for idx, column in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = max(len(str(column)) + 2, 12)
        return

    sized = df.fillna("").astype(str)
    for idx, column in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(column)), *sized[column].map(len).tolist()],
            default=len(str(column)),
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 32)


def apply_number_formats(ws, columns: List[str]) -> None:
    if "Data vencimento" in columns:
        cidx = columns.index("Data vencimento") + 1
        col_letter = get_column_letter(cidx)
        for row in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row}"].number_format = "DD/MM/YYYY"

    for money_col in ["Valor parcela", "Total inadimplente"]:
        if money_col in columns:
            cidx = columns.index(money_col) + 1
            col_letter = get_column_letter(cidx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = 'R$ #,##0.00'


def style_data_sheet(ws, df: pd.DataFrame) -> None:
    columns = list(df.columns)

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.row_dimensions[1].height = 24

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    center_columns = {
        "Seguradora",
        "Dias atraso",
        "Data vencimento",
        "Periodicidade",
        "Método de pagamento",
        "Business Channel",
    }
    right_columns = {"Valor parcela", "Total inadimplente"}

    for row in range(2, ws.max_row + 1):
        fill = ALT_FILL if row % 2 == 0 else None
        for col_idx, column_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER

            if fill is not None:
                cell.fill = fill

            if column_name in center_columns:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif column_name in right_columns:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    apply_number_formats(ws, columns)
    auto_fit_columns(ws, df)

    if ws.max_row >= 2 and "Dias atraso" in columns:
        idx = columns.index("Dias atraso") + 1
        col = get_column_letter(idx)
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="E8F3F1",
                mid_type="percentile",
                mid_value=50,
                mid_color="F3D79A",
                end_type="max",
                end_color="C65146",
            ),
        )

    if ws.max_row >= 2 and "Total inadimplente" in columns:
        idx = columns.index("Total inadimplente") + 1
        col = get_column_letter(idx)
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            ColorScaleRule(
                start_type="min",
                start_color="F7F2E7",
                mid_type="percentile",
                mid_value=50,
                mid_color="DDB774",
                end_type="max",
                end_color="A66B00",
            ),
        )


def write_metric_card(
    ws,
    start_row: int,
    start_col: int,
    title: str,
    value: str,
    fill: PatternFill,
) -> None:
    end_col = start_col + 2
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=end_col,
    )
    ws.merge_cells(
        start_row=start_row + 1,
        start_column=start_col,
        end_row=start_row + 2,
        end_column=end_col,
    )

    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = title
    title_cell.fill = fill
    title_cell.font = CARD_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER

    value_cell = ws.cell(row=start_row + 1, column=start_col)
    value_cell.value = value
    value_cell.fill = fill
    value_cell.font = CARD_VALUE_FONT
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = THIN_BORDER

    for row in range(start_row, start_row + 3):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = THIN_BORDER


def write_dashboard_table(
    ws,
    title: str,
    df: pd.DataFrame,
    start_row: int,
    start_col: int,
    currency_columns: Optional[List[str]] = None,
) -> Tuple[int, int, int]:
    currency_columns = currency_columns or []

    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = SECTION_FONT
    title_cell.fill = SOFT_FILL
    title_cell.border = THIN_BORDER
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    header_row = start_row + 1
    for col_offset, column_name in enumerate(df.columns, start=0):
        cell = ws.cell(row=header_row, column=start_col + col_offset, value=column_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if df.empty:
        empty_row = header_row + 1
        cell = ws.cell(row=empty_row, column=start_col, value="Sem dados")
        cell.border = THIN_BORDER
        return header_row, empty_row, empty_row

    for row_offset, row_data in enumerate(df.itertuples(index=False), start=header_row + 1):
        fill = ALT_FILL if row_offset % 2 == 0 else None
        for col_offset, value in enumerate(row_data, start=0):
            cell = ws.cell(row=row_offset, column=start_col + col_offset, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill is not None:
                cell.fill = fill

            header_name = df.columns[col_offset]
            if header_name in currency_columns and isinstance(value, (int, float)):
                cell.number_format = 'R$ #,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    return header_row, header_row + 1, header_row + len(df)


def save_excel_with_formats(df: pd.DataFrame, path: Path, sheet_name: str = "Dados") -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        style_data_sheet(ws, df)


def save_final_report_with_dashboard(df: pd.DataFrame, path: Path) -> None:
    dashboard_df = df.copy()
    dashboard_df["Seguradora"] = dashboard_df["Seguradora"].fillna("Não informado")
    dashboard_df["Business Channel"] = dashboard_df["Business Channel"].fillna("Não informado")
    dashboard_df["Proprietário da Oportunidade"] = dashboard_df[
        "Proprietário da Oportunidade"
    ].fillna("Não informado")
    dashboard_df["Total inadimplente"] = pd.to_numeric(
        dashboard_df["Total inadimplente"], errors="coerce"
    ).fillna(0)
    dashboard_df["Dias atraso"] = pd.to_numeric(
        dashboard_df["Dias atraso"], errors="coerce"
    ).fillna(0)

    insurer_summary = (
        dashboard_df.groupby("Seguradora", dropna=False)
        .agg(Apólices=("Número apólice", "count"), Total=("Total inadimplente", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
    )

    owner_summary = (
        dashboard_df.groupby("Proprietário da Oportunidade", dropna=False)
        .agg(Apólices=("Número apólice", "count"), Total=("Total inadimplente", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
        .head(10)
    )

    channel_summary = (
        dashboard_df.groupby("Business Channel", dropna=False)
        .agg(Apólices=("Número apólice", "count"), Total=("Total inadimplente", "sum"))
        .reset_index()
        .sort_values("Total", ascending=False)
        .head(8)
    )

    aging_labels = ["0-30 dias", "31-60 dias", "61-90 dias", "91+ dias"]
    aging_summary = (
        pd.cut(
            dashboard_df["Dias atraso"],
            bins=[-1, 30, 60, 90, float("inf")],
            labels=aging_labels,
        )
        .value_counts(sort=False)
        .rename_axis("Faixa")
        .reset_index(name="Apólices")
    )

    total_cases = len(dashboard_df)
    total_inad = float(dashboard_df["Total inadimplente"].sum())
    avg_ticket = float(dashboard_df["Total inadimplente"].mean()) if total_cases else 0.0
    avg_delay = float(dashboard_df["Dias atraso"].mean()) if total_cases else 0.0

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)
        wb = writer.book
        ws_data = writer.sheets["Dados"]
        style_data_sheet(ws_data, df)

        ws_dash = wb.create_sheet("Dashboard", 0)
        wb.active = 0

        ws_dash.sheet_view.showGridLines = False
        ws_dash.sheet_view.zoomScale = 85
        for col in range(1, 19):
            ws_dash.column_dimensions[get_column_letter(col)].width = 16

        ws_dash["A1"] = "Dashboard de Inadimplência"
        ws_dash["A1"].font = TITLE_FONT
        ws_dash["A2"] = (
            f"Atualizado em {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')} a partir do relatório consolidado."
        )
        ws_dash["A2"].font = SUBTITLE_FONT

        write_metric_card(ws_dash, 4, 1, "Apólices no relatório", f"{total_cases:,}".replace(",", "."), CARD_FILL)
        write_metric_card(ws_dash, 4, 5, "Total inadimplente", format_brl(total_inad), ACCENT_FILL)
        write_metric_card(ws_dash, 4, 9, "Ticket médio", format_brl(avg_ticket), CARD_FILL)
        write_metric_card(ws_dash, 4, 13, "Atraso médio", f"{avg_delay:.0f} dias", ACCENT_FILL)

        write_dashboard_table(
            ws_dash,
            "Inadimplência por Seguradora",
            insurer_summary.rename(columns={"Seguradora": "Seguradora"}),
            start_row=14,
            start_col=1,
            currency_columns=["Total"],
        )
        write_dashboard_table(
            ws_dash,
            "Top Proprietários",
            owner_summary.rename(columns={"Proprietário da Oportunidade": "Proprietário"}),
            start_row=14,
            start_col=5,
            currency_columns=["Total"],
        )
        write_dashboard_table(
            ws_dash,
            "Faixas de Atraso",
            aging_summary,
            start_row=14,
            start_col=9,
        )
        write_dashboard_table(
            ws_dash,
            "Canais com Mais Inadimplência",
            channel_summary.rename(columns={"Business Channel": "Canal"}),
            start_row=14,
            start_col=12,
            currency_columns=["Total"],
        )


def read_sheet(path: Path) -> pd.DataFrame:
    return pd.read_excel(path, header=0, dtype=object)


def select_and_rename(df: pd.DataFrame, config: LayoutConfig) -> pd.DataFrame:
    if not config.keep_letters:
        raise ValueError(
            f"Layout da seguradora '{config.seguradora}' não foi configurado (colunas não informadas)."
        )

    selected: Dict[str, pd.Series] = {}

    for letter in config.keep_letters:
        idx = excel_letter_to_index(letter)
        if idx >= len(df.columns):
            raise ValueError(
                f"A coluna {letter} ({idx + 1}) não existe na planilha para {config.seguradora}."
            )

        target = config.source_to_target[letter]
        selected[target] = df.iloc[:, idx]

    cleaned = pd.DataFrame(selected)

    for col in BASE_COLUMNS:
        if col not in cleaned.columns:
            cleaned[col] = pd.NA

    cleaned["Seguradora"] = config.seguradora
    cleaned["Número apólice"] = cleaned["Número apólice"].apply(normalize_policy_number)
    cleaned["Valor parcela"] = cleaned["Valor parcela"].apply(normalize_brl_number)

    data_venc = pd.to_datetime(cleaned["Data vencimento"], errors="coerce", dayfirst=True)
    cleaned["Data vencimento"] = data_venc.dt.date

    today = pd.Timestamp.now().normalize()
    cleaned["Dias atraso"] = (today - data_venc).dt.days.clip(lower=0).fillna(0).astype(int)

    return cleaned[BASE_COLUMNS]


def detect_insurer(file_name: str) -> Optional[str]:
    normalized = file_name.lower().replace("_", " ").replace("-", " ")
    compact = normalized.replace(" ", "")

    for key in FLOW_ORDER:
        for alias in INSURER_ALIASES.get(key, (key,)):
            if alias.lower().replace(" ", "") in compact:
                return key

    return None


def apply_generic_rules(
    df: pd.DataFrame,
    policy_normalizer: Callable[[object], object],
) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    d = df.copy()

    d["Número apólice"] = d["Número apólice"].apply(policy_normalizer)
    d["Data vencimento"] = pd.to_datetime(d["Data vencimento"], errors="coerce", dayfirst=True)
    d["Valor parcela"] = pd.to_numeric(d["Valor parcela"], errors="coerce")

    d = d[d["Número apólice"].notna() & d["Data vencimento"].notna()].copy()

    if d.empty:
        for col in BASE_COLUMNS:
            if col not in d.columns:
                d[col] = pd.Series(dtype="object")
        return d

    d["_mes"] = d["Data vencimento"].dt.to_period("M")

    premio_por_mes = (
        d.groupby(["Número apólice", "_mes"], as_index=False)["Valor parcela"]
        .sum()
        .rename(columns={"Valor parcela": "Premio"})
    )

    total_por_apolice = (
        premio_por_mes.groupby("Número apólice", as_index=False)["Premio"]
        .sum()
        .rename(columns={"Premio": "Total_inad"})
    )

    d = d.merge(premio_por_mes, on=["Número apólice", "_mes"], how="left")
    d = d.merge(total_por_apolice, on="Número apólice", how="left")

    d["Valor parcela"] = d["Premio"]
    d["Total inadimplente"] = d["Total_inad"]

    d = d.sort_values("Data vencimento", ascending=False).drop_duplicates(
        subset=["Número apólice"], keep="first"
    )

    d["Data vencimento"] = d["Data vencimento"].dt.date
    d = d.drop(columns=["_mes", "Premio", "Total_inad"], errors="ignore")

    return d


def apply_mag_rules(df: pd.DataFrame) -> pd.DataFrame:
    return apply_generic_rules(df, normalize_policy_number)


def apply_metlife_rules(df: pd.DataFrame) -> pd.DataFrame:
    return apply_generic_rules(df, metlife_prefix_policy)


def apply_azos_rules(df: pd.DataFrame) -> pd.DataFrame:
    return apply_generic_rules(df, normalize_policy_number)


def apply_omint_rules(df: pd.DataFrame) -> pd.DataFrame:
    return apply_generic_rules(df, normalize_policy_number)


def apply_prudential_rules(df: pd.DataFrame) -> pd.DataFrame:
    return apply_generic_rules(df, normalize_policy_number)


RULES_BY_SEGURADORA: Dict[str, Callable[[pd.DataFrame], pd.DataFrame]] = {
    "MAG": apply_mag_rules,
    "MetLife": apply_metlife_rules,
    "Azos": apply_azos_rules,
    "Omint": apply_omint_rules,
    "Prudential": apply_prudential_rules,
}


def normalize_lookup_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "Número Apólice": "Número apólice",
        "Numero Apolice": "Número apólice",
        "Numero Apólice": "Número apólice",
        "Email do Proprietário da Oportunidade": "E-mail do Proprietário da Oportunidade",
    }

    existing = {old: new for old, new in rename_map.items() if old in df.columns}
    if existing:
        df = df.rename(columns=existing)

    return df


def collapse_lookup_rows(df: pd.DataFrame) -> LookupLoadResult:
    rows: List[Dict[str, object]] = []
    warnings: List[str] = []

    for policy_number, group in df.groupby("Número apólice", sort=False):
        row: Dict[str, object] = {"Número apólice": policy_number}
        conflicting_cols: List[str] = []

        for col in ENRICH_COLUMNS:
            values = distinct_non_empty_values(group[col])
            if len(values) > 1:
                conflicting_cols.append(col)
                row[col] = pd.NA
            elif values:
                row[col] = values[0]
            else:
                row[col] = pd.NA

        if conflicting_cols:
            warnings.append(
                "Base SF com conflito para a apólice "
                f"{policy_number}: colunas deixadas em branco por divergência: "
                f"{', '.join(conflicting_cols)}."
            )

        rows.append(row)

    collapsed = pd.DataFrame(rows, columns=["Número apólice", *ENRICH_COLUMNS])
    return LookupLoadResult(df=collapsed, warnings=warnings)


def load_lookup_excel(
    path: Path,
    sheet_name: str = BASE_SF_SHEET_NAME,
) -> LookupLoadResult:
    if not path.exists():
        raise FileNotFoundError(f"Arquivo da base SF não encontrado: {path}")

    df = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    df = normalize_lookup_columns(df)

    required = ["Número apólice", *ENRICH_COLUMNS]
    missing = [c for c in required if c not in df.columns]

    if missing:
        raise ValueError(f"Colunas ausentes na base SF: {missing}")

    df = df[required].copy()
    df["Número apólice"] = df["Número apólice"].apply(normalize_policy_number)

    df = df.dropna(subset=["Número apólice"]).reset_index(drop=True)

    if df.empty:
        empty = pd.DataFrame(columns=required)
        return LookupLoadResult(df=empty, warnings=[])

    return collapse_lookup_rows(df)


def enrich_with_lookup_excel(
    df: pd.DataFrame,
    lookup_path: Path,
    sheet_name: str = BASE_SF_SHEET_NAME,
) -> Tuple[pd.DataFrame, List[str]]:
    enriched = df.copy()

    for col in ENRICH_COLUMNS:
        if col not in enriched.columns:
            enriched[col] = pd.NA

    if enriched.empty:
        return enriched, []

    enriched["Número apólice"] = enriched["Número apólice"].apply(normalize_policy_number)
    lookup_result = load_lookup_excel(lookup_path, sheet_name=sheet_name)
    lookup = lookup_result.df

    enriched = enriched.merge(
        lookup,
        on="Número apólice",
        how="left",
        suffixes=("", "_sf"),
    )

    for col in ENRICH_COLUMNS:
        col_sf = f"{col}_sf"
        if col_sf in enriched.columns:
            enriched[col] = enriched[col].fillna(enriched[col_sf])
            enriched = enriched.drop(columns=[col_sf])

    return enriched, lookup_result.warnings


def process_folder(
    folder: Path,
    output: Path,
    recursive: bool = True,
) -> None:
    if not folder.exists():
        raise FileNotFoundError(f"Pasta de entrada não encontrada: {folder}")

    file_iter = folder.rglob("*") if recursive else folder.iterdir()
    files = sorted(
        [p for p in file_iter if p.is_file() and p.suffix.lower() in {".xlsx", ".xlsm"}],
        key=lambda p: p.name.lower(),
    )

    if not files:
        print("Nenhum arquivo .xlsx/.xlsm encontrado na pasta informada.")
        return

    output.mkdir(parents=True, exist_ok=True)
    print(f"Arquivos encontrados: {len(files)}")

    per_insurer: Dict[str, List[pd.DataFrame]] = {k: [] for k in FLOW_ORDER}
    skipped: List[str] = []
    warnings: List[str] = []
    processed_count = 0

    for file in files:
        insurer_key = detect_insurer(file.name)
        if not insurer_key:
            skipped.append(f"{file.name}: seguradora não identificada no nome do arquivo")
            continue

        config = CONFIGS[insurer_key]
        if not is_layout_configured(config):
            skipped.append(
                f"{file.name} [{config.seguradora}]: layout ainda não configurado no script"
            )
            continue

        try:
            original = read_sheet(file)
            cleaned = select_and_rename(original, config)
            per_insurer[insurer_key].append(cleaned)
            processed_count += 1

            out_path = output / f"{file.stem}_limpo.xlsx"
            save_excel_with_formats(cleaned, out_path)
            print(f"OK  - {file.name} [{config.seguradora}] -> {out_path.name}")
        except Exception as exc:
            skipped.append(f"{file.name} [{config.seguradora}]: {repr(exc)}")

    consolidated: List[pd.DataFrame] = []
    for insurer_key in FLOW_ORDER:
        consolidated.extend(per_insurer[insurer_key])

    if consolidated:
        adjusted: List[pd.DataFrame] = []

        for part in consolidated:
            seg = (
                str(part["Seguradora"].iloc[0])
                if "Seguradora" in part.columns and not part.empty
                else ""
            )
            rule = RULES_BY_SEGURADORA.get(seg)
            if rule:
                part = rule(part)
            adjusted.append(part)

        final_df = pd.concat(adjusted, ignore_index=True)

        lookup_path = app_base_dir() / "BASE_SF_INADIMPLENCIA" / BASE_SF_FILE_NAME

        try:
            final_df, lookup_warnings = enrich_with_lookup_excel(
                final_df,
                lookup_path,
                sheet_name=BASE_SF_SHEET_NAME,
            )
            warnings.extend(lookup_warnings)
            print("Enriquecimento com BASE_SF_INADIMPLENCIA concluído com sucesso.")
        except Exception as exc:
            print(f"AVISO - não foi possível enriquecer com a base SF: {repr(exc)}")
            warnings.append(f"Enriquecimento com a base SF não executado: {repr(exc)}")
            for col in ENRICH_COLUMNS:
                if col not in final_df.columns:
                    final_df[col] = pd.NA

        final_df = final_df.reindex(columns=FINAL_COLUMNS)
        final_path = output / "relatorio_final.xlsx"
        save_final_report_with_dashboard(final_df, final_path)
        print(f"\nRelatório consolidado criado: {final_path} (linhas: {len(final_df)})")
    else:
        print("\nNenhuma planilha válida processada para consolidação.")

    if warnings:
        print("\nAvisos:")
        for item in warnings:
            print(f"- {item}")

    if skipped:
        print("\nArquivos ignorados / erros:")
        for item in skipped:
            print(f"- {item}")

    print("\nResumo da execução:")
    print(f"- Arquivos processados com sucesso: {processed_count}")
    print(f"- Arquivos ignorados / com erro: {len(skipped)}")
    print(f"- Avisos não bloqueantes: {len(warnings)}")

    if skipped or warnings:
        print("Execução concluída com pendências.")
    else:
        print("Execução concluída com sucesso.")


def app_base_dir() -> Path:
    od = os.environ.get("OneDriveCommercial") or os.environ.get("OneDrive")
    if od:
        fixed = (
            Path(od)
            / "Financial Planning - Automações - Documentos"
            / "Seguro de Vida - Documentos"
            / "Inadimplencia_Automacao"

        )
        if fixed.exists():
            return fixed

    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent

    return Path(__file__).resolve().parent


def parse_args() -> argparse.Namespace:
    base = app_base_dir()

    parser = argparse.ArgumentParser(
        description="Limpeza e consolidação de planilhas de seguradoras"
    )
    parser.add_argument(
        "--input",
        default=str(base / "ENTRADA"),
        help="Pasta com os arquivos de entrada (padrão: ENTRADA ao lado do app)",
    )
    parser.add_argument(
        "--output",
        default=str(base / "SAIDA"),
        help="Pasta para salvar os arquivos limpos e o relatório final (padrão: SAIDA ao lado do app)",
    )
    parser.add_argument(
        "--no-recursive",
        action="store_true",
        help="Desativa busca em subpastas (por padrão a busca é recursiva)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    process_folder(
        Path(args.input),
        Path(args.output),
        recursive=not args.no_recursive,
    )


if __name__ == "__main__":
    main()
